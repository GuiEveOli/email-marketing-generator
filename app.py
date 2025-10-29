# app.py

import time
import re
import csv
import os
import io
import shutil
import tempfile
import uuid
import unicodedata
from urllib.parse import quote_plus
from bs4 import BeautifulSoup
from flask import Flask, render_template, request, send_file, jsonify, redirect, url_for, send_from_directory
from datetime import datetime

# Importa utilitários do Google Sheets
ler_produtos_via_sheets_csv = None
ler_produtos_via_api = None
try:
    from google_sheets import ler_produtos_via_sheets_csv, ler_produtos_via_api
    print("✓ Módulo google_sheets carregado com sucesso")
    print(f"  - ler_produtos_via_sheets_csv: {ler_produtos_via_sheets_csv}")
    print(f"  - ler_produtos_via_api: {ler_produtos_via_api}")
except ImportError as e:
    print(f"⚠ google_sheets não encontrado: {e}")
except Exception as e:
    print(f"⚠ Erro ao importar google_sheets: {e}")
    import traceback
    traceback.print_exc()

# --- Configuração do Flask ---
app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False  # Permite caracteres não-ASCII no JSON
app.config['JSON_SORT_KEYS'] = False  # Mantém ordem das chaves

# Torna 'current' disponível em todos os templates (ex.: para destacar item do menu)
@app.context_processor
def inject_current():
    return {'current': request.path}

# Cache para produtos (evita recarregar CSV toda vez)
_produtos_cache = None
_cache_timestamp = None

# --- FUNÇÃO PARA CARREGAR PRODUTOS DA PLANILHA ---
def carregar_produtos_planilha(force_reload=False, csv_filename: str | None = None):
    """
    Carrega os produtos da fonte configurada:
    - Google Sheets (CSV export ou API) se DATA_SOURCE estiver definido
    - Fallback: CSV local
    """
    global _produtos_cache, _cache_timestamp

    if not force_reload and _produtos_cache is not None:
        return _produtos_cache

    produtos = []

    # 1) Tenta Google Sheets conforme DATA_SOURCE
    data_source = (os.getenv('DATA_SOURCE') or '').strip().lower()
    try:
        if data_source == 'sheets_csv' and ler_produtos_via_sheets_csv:
            sheets_url = os.getenv('SHEETS_CSV_URL', '').strip()
            if sheets_url:
                print("→ Lendo produtos via Google Sheets (CSV export)...")
                produtos, meta = ler_produtos_via_sheets_csv(sheets_url)
                if produtos:
                    print(f"✓ Sheets CSV carregado. Total: {len(produtos)} | Delimitador: {meta.get('delimiter')}")
                    _produtos_cache = produtos
                    _cache_timestamp = time.time()
                    return produtos
                else:
                    print("⚠ Sheets CSV retornou vazio. Indo para fallback (CSV local).")
            else:
                print("⚠ SHEETS_CSV_URL não configurado. Indo para fallback (CSV local).")

        if data_source == 'sheets_api' and ler_produtos_via_api:
            sheets_id = os.getenv('SHEETS_ID', '').strip()
            sheets_tab_or_range = os.getenv('SHEETS_TAB', '').strip() or os.getenv('SHEETS_RANGE', '').strip()
            creds_path = os.getenv('SHEETS_CREDENTIALS_FILE', '').strip()
            if sheets_id and creds_path:
                print("→ Lendo produtos via Google Sheets API...")
                produtos, meta = ler_produtos_via_api(sheets_id, sheets_tab_or_range or None, creds_path)
                if produtos:
                    print(f"✓ Sheets API carregado. Total: {len(produtos)} | Worksheet: {meta.get('worksheet')}")
                    _produtos_cache = produtos
                    _cache_timestamp = time.time()
                    return produtos
                else:
                    print(f"⚠ Sheets API sem dados válidos. Headers: {meta.get('headers')} | Fallback (CSV local).")
            else:
                print("⚠ SHEETS_ID ou SHEETS_CREDENTIALS_FILE ausente. Fallback (CSV local).")
    except Exception as e:
        print(f"✗ Erro ao ler do Google Sheets ({data_source}): {e}. Fallback (CSV local).")

    # 2) Fallback para CSV local (com autodetecção)
    # Se já tem cache e não forçou reload, retorna do cache
    if not force_reload and _produtos_cache is not None:
        return _produtos_cache
    
    produtos = []
    
    # Permite sobrescrever via argumento ou variável de ambiente
    csv_candidates = []
    if csv_filename:
        csv_candidates.append(csv_filename)
    env_csv = os.getenv('PRODUTOS_CSV')
    if env_csv:
        csv_candidates.append(env_csv)
    # Ordem de preferência padrão
    csv_candidates.extend(['produtos.csv', 'produtos2.csv'])

    base_dir = os.path.dirname(__file__)
    csv_path = None
    for candidate in csv_candidates:
        candidate_path = candidate if os.path.isabs(candidate) else os.path.join(base_dir, candidate)
        if os.path.exists(candidate_path):
            csv_path = candidate_path
            break

    if not csv_path:
        print("✗ Nenhum arquivo de produtos encontrado.")
        print("  Procurei por: ", ', '.join(csv_candidates))
        return produtos
    
    # Lista de codificações para tentar
    encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'iso-8859-1', 'cp1252', 'windows-1252']
    # Lista de delimitadores para tentar
    delimiters = [';', ',', '\t', '|']
    
    def _norm(s: str | None) -> str:
        if s is None:
            return ''
        # Remove BOM, espaços e normaliza para UPPER sem underscores diferentes
        return re.sub(r"\s+", '', str(s).replace('\ufeff', '').strip()).upper()

    expected = ['COD_PRODUTO', 'NOMEPRODUTOECOMM', 'COD_BARRAS', 'URLECOMMERCEIMG', 'PRODUCTURL']
    expected_norm = [_norm(x) for x in expected]
    # Aliases comuns para tolerar planilhas variantes
    aliases = {
        'CODPRODUTO': 'COD_PRODUTO',
        'SKU': 'COD_PRODUTO',
        'NOME': 'NOMEPRODUTOECOMM',
        'NOMEPRODUTO': 'NOMEPRODUTOECOMM',
        'NOME_PRODUTO': 'NOMEPRODUTOECOMM',
        'DESCRICAO': 'NOMEPRODUTOECOMM',
        'CODBARRAS': 'COD_BARRAS',
        'EAN': 'COD_BARRAS',
        'URLIMG': 'URLECOMMERCEIMG',
        'IMAGEM': 'URLECOMMERCEIMG',
        'URLIMAGEM': 'URLECOMMERCEIMG',
        'URL': 'PRODUCTURL',
        'LINK': 'PRODUCTURL',
        'PRODUCT_URL': 'PRODUCTURL',
        'URL_PRODUTO': 'PRODUCTURL'
    }

    last_detected_headers = None
    
    for encoding in encodings:
        for delimiter in delimiters:
            try:
                with open(csv_path, 'r', encoding=encoding, errors='replace', newline='') as file:
                    reader = csv.DictReader(file, delimiter=delimiter)
                    fieldnames = reader.fieldnames or []
                    detected_norm = [_norm(h) for h in fieldnames]
                    last_detected_headers = fieldnames[:]

                    # Monta um mapeamento de nome esperado -> nome real da planilha
                    mapping: dict[str, str] = {}
                    for idx, norm_name in enumerate(detected_norm):
                        real_name = fieldnames[idx]
                        # Se for exatamente um dos esperados
                        if norm_name in expected_norm:
                            i = expected_norm.index(norm_name)
                            mapping[expected[i]] = real_name
                        # Se for um alias, mapeia para o esperado correspondente
                        elif norm_name in aliases:
                            canonical = aliases[norm_name]
                            # Mapeia apenas se ainda não definido
                            mapping.setdefault(canonical, real_name)

                    # Verifica se conseguimos mapear todos os esperados
                    if all(col in mapping for col in expected):
                        rows = list(reader)
                        produtos_lidos = []
                        for row in rows:
                            try:
                                sku = str(row.get(mapping['COD_PRODUTO'], '')).strip()
                                nome = str(row.get(mapping['NOMEPRODUTOECOMM'], '')).strip()
                                ean = str(row.get(mapping['COD_BARRAS'], '')).strip()
                                imagem = str(row.get(mapping['URLECOMMERCEIMG'], '')).strip()
                                url = str(row.get(mapping['PRODUCTURL'], '')).strip()

                                # Ignora linhas claramente inválidas (sem URL ou com marcador 0/-)
                                if not url or url in ('0', '-'):
                                    continue

                                produtos_lidos.append({
                                    'sku': sku,
                                    'nome': nome,
                                    'ean': ean,
                                    'imagem': imagem,
                                    'url': url
                                })
                            except Exception:
                                # Pula linha problemática, segue nas demais
                                continue

                        if produtos_lidos:
                            produtos = produtos_lidos
                            print("✓ CSV carregado com sucesso!")
                            print(f"  - Arquivo: {os.path.basename(csv_path)}")
                            print(f"  - Encoding: {encoding}")
                            print(f"  - Delimitador: '{delimiter}'")
                            print(f"  - Total de produtos válidos: {len(produtos)}")
                            if produtos:
                                print(f"  - Exemplo: {produtos[0]['nome'][:50]}...")

                            _produtos_cache = produtos
                            _cache_timestamp = time.time()
                            return produtos
                        else:
                            # Mesmo com cabeçalhos válidos, não havia linhas úteis
                            continue

            except UnicodeDecodeError:
                continue
            except Exception as e:
                # Mantém a estratégia de continuar tentando, mas com um pouco de diagnóstico
                continue

    print("✗ Não foi possível ler o arquivo CSV.")
    print("  Verifique se as colunas estão corretas (e.g.:)")
    print("  COD_PRODUTO, NOMEPRODUTOECOMM, COD_BARRAS, URLECOMMERCEIMG, PRODUCTURL")
    if last_detected_headers is not None:
        print("  Cabeçalhos detectados na última tentativa:")
        print("  ", last_detected_headers)
    print(f"  Arquivo tentado: {csv_path}")
    return produtos

# --- FUNÇÃO PARA BUSCAR PRODUTO POR SKU, EAN OU URL ---
def buscar_produto(termo_busca):
    """
    Busca produto por SKU (COD_PRODUTO), EAN (COD_BARRAS) ou URL na planilha
    """
    produtos = carregar_produtos_planilha()
    termo_busca = termo_busca.strip().lower()
    
    for produto in produtos:
        if (produto['sku'].lower() == termo_busca or 
            produto['ean'].lower() == termo_busca or
            produto['url'].lower() == termo_busca or
            termo_busca in produto['nome'].lower()):
            return produto
    
    return None

# --- FUNÇÃO AUXILIAR PARA ADICIONAR UTM ---
def adicionar_utm_na_url(url_original, utm_source, utm_medium, utm_campaign):
    """
    Adiciona os parâmetros UTM à URL do produto.
    """
    utm_source_encoded = quote_plus(utm_source)
    utm_medium_encoded = quote_plus(utm_medium)
    utm_campaign_encoded = quote_plus(utm_campaign)
    
    separador = '&' if '?' in url_original else '?'
    utm_params = f"utm_source={utm_source_encoded}&utm_medium={utm_medium_encoded}&utm_campaign={utm_campaign_encoded}"
    return f"{url_original}{separador}{utm_params}"

# --- LÓGICA DE BUSCA DE PRODUTOS ---
def buscar_produtos(produtos_info, template_base_html, utm_source="email-mkt", utm_campaign="cupom+15+novo+site", cor_botao="#ff0000"):
    """
    Busca dados do produto via requests, extraindo o JSON do window.APOLLO_STATE
    contido na tag <script id="main-states"> e gerando os blocos HTML.
    A imagem do produto é preferencialmente obtida do HTML, com fallback no APOLLO_STATE.
    """
    import json
    import requests
    from bs4 import BeautifulSoup

    def _extract_apollo_state(script_text: str) -> dict | None:
        if not script_text:
            return None
        idx = script_text.find("window.APOLLO_STATE")
        if idx == -1:
            return None
        start = script_text.find("{", idx)
        if start == -1:
            return None
        braces = 0
        end = -1
        for i in range(start, len(script_text)):
            ch = script_text[i]
            if ch == "{":
                braces += 1
            elif ch == "}":
                braces -= 1
                if braces == 0:
                    end = i + 1
                    break
        if end == -1:
            return None
        json_str = script_text[start:end]
        json_str = json_str.replace(": undefined", ": null").replace(":undefined", ": null")
        try:
            return json.loads(json_str)
        except Exception:
            try:
                return json.loads(json_str.encode("utf-8", "ignore").decode("utf-8"))
            except Exception:
                return None

    def _to_float(val) -> float:
        try:
            if isinstance(val, (int, float)):
                return float(val)
            if isinstance(val, str):
                v = val.strip()
                if v.count(",") == 1 and v.count(".") == 0:
                    v = v.replace(".", "").replace(",", ".")
                else:
                    v = re.sub(r"[^0-9\.,-]", "", v)
                    if v.count(",") == 1 and v.count(".") == 0:
                        v = v.replace(",", ".")
                return float(v)
        except Exception:
            return 0.0
        return 0.0

    def _format_brl(num: float) -> str:
        return f"{num:.2f}".replace(".", ",")

    def _resolve_image_url(apollo: dict, image_ref) -> str:
        """
        Fallback: resolve a URL via APOLLO_STATE (File:... -> url)
        """
        placeholder = "https://via.placeholder.com/120"
        ref_key = None
        if isinstance(image_ref, str):
            ref_key = image_ref
        elif isinstance(image_ref, dict):
            ref_key = image_ref.get("__ref") or image_ref.get("id")
        if not ref_key:
            return placeholder
        file_obj = apollo.get(ref_key)
        if isinstance(file_obj, dict):
            return file_obj.get("url") or file_obj.get("src") or placeholder
        return placeholder

    def _extract_image_from_html(soup: BeautifulSoup, nome_produto: str | None = None) -> str:
        """
        Tenta obter a imagem diretamente do HTML (galeria/og:image), sem depender do APOLLO_STATE.
        """
        placeholder = "https://via.placeholder.com/120"

        # Preferir a imagem ativa da galeria do produto
        selectors = [
            'div.product-image-gallery-active-image img[src]',
            'div.product-image-gallery-desktop-view img[src]',
            'div.product-image-gallery img[src]',
            'div.static-image-viewer-container img[src]',
            'img[class*="product-image"][src]',
            'img[src][alt]'
        ]
        candidates = []
        for sel in selectors:
            el = soup.select_one(sel)
            if el and el.get('src'):
                src = (el.get('src') or '').strip()
                if src and not src.startswith('data:'):
                    candidates.append((src, el.get('alt') or ''))

        # Se achou, preferir o que casa com o nome do produto
        if candidates:
            if nome_produto:
                for src, alt in candidates:
                    if nome_produto.lower() in alt.lower():
                        return src
            return candidates[0][0]

        # Meta og:image como fallback
        og = soup.find('meta', property='og:image') or soup.find('meta', attrs={'name': 'og:image'})
        if og and og.get('content'):
            content = og.get('content').strip()
            if content:
                return content

        # link rel="image_src"
        link_img = soup.find('link', rel='image_src')
        if link_img and link_img.get('href'):
            href = link_img.get('href').strip()
            if href:
                return href

        # Último recurso: qualquer <img src>
        any_img = soup.find('img', src=True)
        if any_img and any_img.get('src'):
            return any_img.get('src').strip()

        return placeholder

    def _pick_product(apollo: dict) -> tuple[dict | None, str | None]:
        for key, value in apollo.items():
            if isinstance(key, str) and key.startswith("PublicViewerProduct:") and isinstance(value, dict):
                product_id = key.split(":", 1)[1]
                return value, product_id
        return None, None

    def _pick_pricing(apollo: dict, product_id: str | None) -> dict | None:
        candidates = []
        for key, value in apollo.items():
            if isinstance(key, str) and key.startswith("PublicViewerProductPricing:") and isinstance(value, dict):
                if product_id and product_id in key:
                    candidates.append(value)
        if not candidates:
            for key, value in apollo.items():
                if isinstance(key, str) and key.startswith("PublicViewerProductPricing:") and isinstance(value, dict):
                    if "price" in value or "promotionalPrice" in value:
                        candidates.append(value)
        candidates.sort(key=lambda v: ("promotionalPrice" in v, "price" in v), reverse=True)
        return candidates[0] if candidates else None

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
        "Connection": "keep-alive",
    })

    todos_os_produtos_html = []
    contador_produto = 0

    for produto_info in produtos_info:
        url = (produto_info.get('url') or '').strip()
        is_clube = bool(produto_info.get('is_clube', False))
        is_exclusivo = bool(produto_info.get('is_exclusivo', False))
        is_oferta_relampago = bool(produto_info.get('is_oferta_relampago', False))

        if not url:
            continue

        contador_produto += 1
        if contador_produto > 50:
            print("Limite de 50 produtos atingido.")
            break

        try:
            resp = session.get(url, timeout=(5, 12))
            resp.raise_for_status()
        except Exception as e:
            print(f"✗ Falha ao baixar HTML do produto {contador_produto}: {e}")
            continue

        soup = BeautifulSoup(resp.text, 'html.parser')
        script_tag = soup.find('script', id='main-states')

        if not script_tag:
            print(f"✗ <script id='main-states'> não encontrado na página {url}")
            continue

        apollo = _extract_apollo_state(script_tag.text)
        if not apollo:
            print(f"✗ Não foi possível extrair/parsing do window.APOLLO_STATE para {url}")
            continue

        product_obj, product_id = _pick_product(apollo)
        if not product_obj:
            print(f"✗ PublicViewerProduct não encontrado em APOLLO_STATE para {url}")
            continue

        # Nome
        nome_produto = product_obj.get("name") or "Produto Genérico"

        # Imagem: agora prioriza HTML; se falhar, usa APOLLO_STATE
        url_imagem = _extract_image_from_html(soup, nome_produto)
        if not url_imagem or url_imagem.endswith(('.svg', '.gif')):  # pequeno filtro defensivo
            url_imagem = _resolve_image_url(apollo, product_obj.get("image"))

        # Pricing
        pricing = _pick_pricing(apollo, product_id)
        preco_de_num = _to_float(pricing.get("price") if pricing else 0)
        promocional = pricing.get("promotionalPrice") if pricing else None
        preco_por_num = _to_float(promocional if promocional not in (None, "") else 0)

        # DEBUG: Tenta capturar preço promocional de outras fontes se não encontrou no APOLLO_STATE
        if preco_por_num == 0 and preco_de_num > 0:
            print(f"  [DEBUG] Preço promocional não encontrado no APOLLO_STATE, tentando outras fontes...")
            
            # Tenta pegar do HTML diretamente - várias possíveis classes/estruturas
            selectors_preco = [
                'span.promotion',
                'span[class*="promotion"]',
                'span[class*="promotional"]',
                'div[class*="price"] span[class*="promotional"]',
                'div[class*="price"] span[class*="promotion"]',
                'span[class*="sale"]',
                'span[class*="discount"]',
                '.product-price-promotional',
                '.promotional-price',
                '.sale-price',
                'div.product-renderer-pricing-wrapper span[class*="promotion"]',
                'div.product-renderer-active-price-wrapper span',
            ]
            
            for selector in selectors_preco:
                elementos = soup.select(selector)
                for el in elementos:
                    texto = el.get_text(strip=True)
                    print(f"    Tentando selector '{selector}': {texto}")
                    # Tenta extrair número do texto
                    match = re.search(r'R?\$?\s*(\d+[.,]\d{2})', texto)
                    if match:
                        valor_encontrado = _to_float(match.group(1))
                        if valor_encontrado > 0 and valor_encontrado < preco_de_num:
                            print(f"    ✓ Preço promocional encontrado via HTML: R$ {_format_brl(valor_encontrado)}")
                            preco_por_num = valor_encontrado
                            break
                if preco_por_num > 0:
                    break
            
            # Se ainda não encontrou, tenta procurar no APOLLO_STATE completo por qualquer campo com "promotion" ou "sale"
            if preco_por_num == 0:
                print(f"    Vasculhando APOLLO_STATE completo...")
                for key, value in apollo.items():
                    if isinstance(value, dict):
                        for field, field_value in value.items():
                            if 'promotion' in field.lower() or 'sale' in field.lower() or 'discount' in field.lower():
                                tentativa = _to_float(field_value)
                                if tentativa > 0 and tentativa < preco_de_num:
                                    print(f"    ✓ Encontrado em APOLLO_STATE[{key}][{field}]: R$ {_format_brl(tentativa)}")
                                    preco_por_num = tentativa
                                    break
                    if preco_por_num > 0:
                        break

        # Se não encontrou preço promocional, mantém o preço POR zerado para ajuste manual
        # mas garante que o preço DE tenha valor
        if preco_por_num == 0 and preco_de_num == 0:
            # Se ambos estão zerados, não há dados de preço
            preco_de_num = 0
            preco_por_num = 0
        elif preco_por_num == 0 and preco_de_num > 0:
            # Preço DE existe, mas POR está zerado - mantém assim para ajuste manual
            print(f"  ⚠ Preço promocional não encontrado em nenhuma fonte. Mantendo zerado para ajuste manual.")
            pass
        elif preco_de_num == 0 and preco_por_num > 0:
            # Só tem preço promocional (caso raro), usa como preço DE também
            preco_de_num = preco_por_num

        # Cálculo de desconto
        porcentagem_desconto = 0
        if preco_de_num > preco_por_num and preco_por_num > 0:
            try:
                porcentagem_desconto = int(((preco_de_num - preco_por_num) / max(preco_de_num, 0.0001)) * 100)
            except Exception:
                porcentagem_desconto = 0

        preco_por_formatado = _format_brl(preco_por_num)
        preco_de_formatado = _format_brl(preco_de_num)

        utm_medium_automatico = f"produto {contador_produto:02d}"
        url_com_utm = adicionar_utm_na_url(url, utm_source, utm_medium_automatico, utm_campaign)

        # LOG DE CONFERÊNCIA DO PRODUTO
        try:
            print("-" * 60)
            print(f"[Produto {contador_produto:02d}] {nome_produto}")
            print(f"  - ID: {product_id}")
            print(f"  - Preço DE: R$ {preco_de_formatado} | POR: R$ {preco_por_formatado} | Desc.: {porcentagem_desconto}%")
            if is_clube or is_exclusivo or is_oferta_relampago:
                selos = ' | '.join(filter(None, [
                    "Clube" if is_clube else None, 
                    "Exclusivo Site" if is_exclusivo else None,
                    "Oferta Relâmpago" if is_oferta_relampago else None
                ]))
                if selos:
                    print(f"  - Selos: {selos}")
            print(f"  - Imagem (HTML→fallback APOLLO): {url_imagem}")
            print(f"  - URL: {url}")
            print(f"  - URL (UTM): {url_com_utm}")
        except Exception:
            pass

        # Selos - Permite múltiplos selos simultaneamente
        selos_spans = []
        
        if is_clube:
            selos_spans.append('<span style="background-color: #cce0ff; color: #034abb; padding: 4px 8px; border-radius: 6px; font-size: 12px; font-weight: bold; font-family: \'Roboto\', Arial, sans-serif; margin-right: 4px; display: inline-block;">Clube</span>')
        
        if is_exclusivo:
            selos_spans.append('<span style="background-color: #bccdee; color: #122447; padding: 4px 8px; border-radius: 6px; font-size: 12px; font-weight: bold; font-family: \'Roboto\', Arial, sans-serif; margin-right: 4px; display: inline-block;">Exclusivo Site</span>')
        
        if is_oferta_relampago:
            selos_spans.append('<span style="background-color: #ffd700; color: #000000; padding: 4px 8px; border-radius: 6px; font-size: 12px; font-weight: bold; font-family: \'Roboto\', Arial, sans-serif; margin-right: 4px; display: inline-block;">Oferta Relâmpago</span>')
        
        # Se não tiver nenhum selo especial, mostra "Oferta" se houver desconto
        if not selos_spans and porcentagem_desconto > 0:
            selos_spans.append('<span style="background-color: #ffebee; color: #dc3545; padding: 4px 8px; border-radius: 6px; font-size: 12px; font-weight: bold; font-family: \'Roboto\', Arial, sans-serif; margin-right: 4px; display: inline-block;">Oferta</span>')
        
        # Monta o HTML dos selos em uma única linha
        html_selo_oferta = ''
        if selos_spans:
            html_selo_oferta = f'<tr><td align="left" valign="top" style="padding-bottom: 8px; line-height: 1.5;">{"".join(selos_spans)}</td></tr>'

        html_bloco_desconto = ""
        # Mostra o bloco de desconto se houver desconto real
        if porcentagem_desconto > 0 and preco_de_num > 0:
            html_bloco_desconto = f'<tr><td style="padding-bottom: 4px; text-align:left;"><table class="price-table" border="0" cellpadding="0" cellspacing="0" style="width:auto; margin:0;"><tbody><tr><td align="left" valign="middle" style="white-space:nowrap;"><span style="text-decoration: line-through; color: #6c757d; font-size: 12px; font-family: \'Roboto\', Arial, sans-serif;">R$ {preco_de_formatado}</span></td><td align="left" valign="middle" style="padding-left: 10px; white-space:nowrap;"><span style="background-color: #ffebee; color: #dc3545; padding: 4px 8px; border-radius: 6px; font-size: 12px; font-weight: bold; font-family: \'Roboto\', Arial, sans-serif;">-{porcentagem_desconto}%</span></td></tr></tbody></table></td></tr>'
        # Se não tem desconto mas tem preço DE (para mostrar que o preço POR está zerado e precisa ajuste)
        elif preco_de_num > 0 and preco_por_num == 0:
            html_bloco_desconto = f'<tr><td style="padding-bottom: 4px; text-align:left;"><table class="price-table" border="0" cellpadding="0" cellspacing="0" style="width:auto; margin:0;"><tbody><tr><td align="left" valign="middle" style="white-space:nowrap;"><span style="text-decoration: line-through; color: #6c757d; font-size: 12px; font-family: \'Roboto\', Arial, sans-serif;">R$ {preco_de_formatado}</span></td><td align="left" valign="middle" style="padding-left: 10px; white-space:nowrap;"><span style="background-color: #ffebee; color: #dc3545; padding: 4px 8px; border-radius: 6px; font-size: 12px; font-weight: bold; font-family: \'Roboto\', Arial, sans-serif;">-{porcentagem_desconto}%</span></td></tr></tbody></table></td></tr>'

        template_produto = f"""
<!-- Início | Produto -->
<div class="column" style="display: inline-block; width: 50%; max-width: 300px; vertical-align: top; box-sizing: border-box; padding: 4px;">
    <table class="product-card-table" width="100%" border="0" cellpadding="0" cellspacing="0" 
           style="background-color: #ffffff; border-radius: 16px; padding: 12px; text-align: left; height: 172px; box-sizing: border-box;">
        <tbody>
            <tr>
                <!-- Coluna da Imagem -->
                <td class="product-image-cell" valign="top" align="center" style="width: 120px;">
                    <table width="100%" border="0" cellpadding="0" cellspacing="0">
                        <tbody>
                            {html_selo_oferta}
                            <tr>
                                <td align="center" valign="top">
                                    <a target="_blank" href="{url_com_utm}">
                                        <img alt="{nome_produto}" 
                                             style="display: block; margin: 0px auto; max-width: 120px;" 
                                             src="{url_imagem}" />
                                    </a>
                                </td>
                            </tr>
                        </tbody>
                    </table>
                </td>
                
                <!-- Coluna das Informações -->
                <td class="product-info-cell" valign="top" align="left" 
                    style="text-align:left; padding:12px 0 0 12px;">
                    <table width="100%" border="0" cellpadding="0" cellspacing="0">
                        <tbody>
                            <!-- Nome do Produto -->
                            <tr>
                                <td style="font-size: 12px; font-weight: 700; color: #212529; 
                                           font-family: 'Roboto', Arial, sans-serif; padding-bottom: 12px; 
                                           height: 48px; vertical-align: top;">
                                    {nome_produto}
                                </td>
                            </tr>
                            
                            <!-- Bloco de Desconto (se houver) -->
                            {html_bloco_desconto}
                            
                            <!-- Preço -->
                            <tr>
                                <td style="font-size: 16px; font-weight: 700; color: #212529; 
                                           font-family: 'Roboto', Arial, sans-serif; padding-bottom: 12px;">
                                    R$ {_format_brl(preco_por_num)}
                                </td>
                            </tr>
                            
                            <!-- Botão Ver Produto -->
                            <tr>
                                <td>
                                    <a target="_blank" 
                                       style="background-color:{cor_botao};
                                              border-radius:50px;
                                              color:#ffffff;
                                              display:block;
                                              font-family:'Roboto', Arial, sans-serif;
                                              font-size:12px;
                                              font-weight:bold;
                                              height:28px;
                                              line-height:28px;
                                              text-align:center;
                                              text-decoration:none;
                                              width:100%;
                                              -webkit-text-size-adjust:none;" 
                                       href="{url_com_utm}">
                                        Ver Produto
                                    </a>
                                </td>
                            </tr>
                        </tbody>
                    </table>
                </td>
            </tr>
        </tbody>
    </table>
</div>
<!-- Fim | Produto -->
"""
        todos_os_produtos_html.append(template_produto)
        print(f"✓ Produto {contador_produto} processado (via HTML/APOLLO).")

    html_final_dos_produtos = '\n'.join(todos_os_produtos_html)

    if '{{PRODUTOS_PLACEHOLDER}}' in template_base_html:
        email_final_html = template_base_html.replace('{{PRODUTOS_PLACEHOLDER}}', html_final_dos_produtos)
    elif '<!-- PRODUTOS -->' in template_base_html:
        email_final_html = template_base_html.replace('<!-- PRODUTOS -->', html_final_dos_produtos)
    elif '<!-- PRODUTOS_AQUI -->' in template_base_html:
        email_final_html = template_base_html.replace('<!-- PRODUTOS_AQUI -->', html_final_dos_produtos)
    else:
        email_final_html = template_base_html + html_final_dos_produtos

    return email_final_html

# --- ROTAS DO SITE ---

@app.route('/')
def index():
    """Página inicial - Hub de ferramentas"""
    return render_template('index.html')

@app.route('/gerador')
def gerador():
    """Página do gerador de email marketing"""
    return render_template('gerador.html')

@app.route('/skuconsult')
def skuconsult():
    """Página de consulta de SKU"""
    return render_template('skuconsult/index.html')

@app.route('/organizador')
def organizador():
    """Página do organizador de pastas"""
    return render_template('organizador.html')

@app.route('/buscar-sugestoes', methods=['POST'])
def buscar_sugestoes():
    """
    API para buscar sugestões de produtos - busca 100% EXATA
    Só mostra produtos que correspondem EXATAMENTE ao termo digitado
    """
    data = request.get_json()
    termo_busca = data.get('termo', '').strip()
    
    if not termo_busca:
        return jsonify({
            'success': True,
            'sugestoes': []
        })
    
    # Se for numérico, aceita 1+ caracteres, senão 2+
    is_numerico = termo_busca.isdigit()
    min_chars = 1 if is_numerico else 2
    
    if len(termo_busca) < min_chars:
        return jsonify({
            'success': True,
            'sugestoes': []
        })
    
    produtos = carregar_produtos_planilha()
    sugestoes = []
    termo_lower = termo_busca.lower()
    
    for produto in produtos:
        # Busca 100% EXATA por SKU ou EAN - só mostra se for idêntico
        if (produto['sku'].lower() == termo_lower or 
            produto['ean'].lower() == termo_lower):
            sugestoes.append(produto)
            
        # Limita a 10 sugestões
        if len(sugestoes) >= 10:
            break
    
    return jsonify({
        'success': True,
        'sugestoes': sugestoes
    }), 200, {'Content-Type': 'application/json; charset=utf-8'}

@app.route('/buscar-produto', methods=['POST'])
def buscar_produto_api():
    """
    API para buscar produto por SKU, EAN ou URL (busca exata)
    """
    data = request.get_json()
    termo_busca = data.get('termo', '').strip()
    
    if not termo_busca:
        return jsonify({'error': 'Termo de busca vazio'}), 400
    
    # Tenta buscar na planilha - busca exata
    produtos = carregar_produtos_planilha()
    termo_lower = termo_busca.lower()
    
    for produto in produtos:
        if (produto['sku'].lower() == termo_lower or 
            produto['ean'].lower() == termo_lower or
            produto['url'].lower() == termo_lower):
            return jsonify({
                'success': True,
                'produto': produto,
                'fonte': 'planilha'
            }), 200, {'Content-Type': 'application/json; charset=utf-8'}
    
    # Se não encontrar na planilha e for uma URL válida
    if termo_busca.startswith('http'):
        return jsonify({
            'success': True,
            'produto': {
                'url': termo_busca,
                'nome': 'Produto via URL',
                'imagem': 'https://via.placeholder.com/120',
                'sku': 'URL',
                'ean': '-'
            },
            'fonte': 'url'
        }), 200, {'Content-Type': 'application/json; charset=utf-8'}
    
    return jsonify({
        'success': False,
        'error': 'Produto não encontrado'
    }), 404

@app.route('/gerar', methods=['POST'])
def gerar_email():
    try:
        data = request.get_json()
        
        if not data:
            print("ERRO: Nenhum dado recebido no request")
            return jsonify({
                'success': False,
                'error': 'Nenhum dado recebido'
            }), 400
        
        produtos_selecionados = data.get('produtos', [])
        
        if not produtos_selecionados:
            print("ERRO: Nenhum produto selecionado")
            return jsonify({
                'success': False,
                'error': 'Nenhum produto selecionado'
            }), 400
        
        utm_source = data.get('utm_source', 'email-mkt')
        utm_campaign = data.get('utm_campaign', 'sem-campanha')
        cta_url = data.get('cta_url', 'https://www.superkoch.com.br/promocoes')
        preheader_text = data.get('preheader_text', 'Peça até as 15h e receba HOJE. Sujeito a disponibilidade.')
        apenas_produtos = data.get('apenas_produtos', '') == 'true'
        bloco_03_selecionado = data.get('componente_bloco_03')
        bloco_05_selecionado = data.get('componente_bloco_05')
        bloco_cupom_selecionado = data.get('componente_bloco_cupom', '')
        cor_botao = data.get('cor_botao', '#ff0000')
        
        # Se "apenas produtos" estiver ativo, limpa os outros componentes
        if apenas_produtos:
            bloco_03_selecionado = None
            bloco_05_selecionado = None
            bloco_cupom_selecionado = ''
        
        # Valida formato hexadecimal
        if not re.match(r'^#[0-9A-Fa-f]{6}$', cor_botao):
            return jsonify({
                'success': False,
                'error': 'Cor do botão inválida. Use formato hexadecimal (#RRGGBB)'
            }), 400

        print(f"✓ Recebidos {len(produtos_selecionados)} produtos para processar.")
        print(f"✓ UTM Source: {utm_source}")
        print(f"✓ UTM Campaign: {utm_campaign}")
        print(f"✓ CTA URL: {cta_url}")
        print(f"✓ Preheader: {preheader_text}")
        print(f"✓ Apenas Produtos: {'Sim' if apenas_produtos else 'Não'}")
        print(f"✓ Bloco Cupom: {'Sim' if bloco_cupom_selecionado else 'Não selecionado'}")
        print(f"✓ Cor do botão: {cor_botao}")
        
        # Adiciona UTM ao CTA
        cta_url_com_utm = adicionar_utm_na_url(cta_url, utm_source, "todas as ofertas", utm_campaign)
        
        print("→ Renderizando template base...")
        template_para_produtos = render_template(
            'email_layout.html', 
            componente_bloco_03=bloco_03_selecionado, 
            componente_bloco_05=bloco_05_selecionado,
            componente_bloco_cupom=bloco_cupom_selecionado,
            cta_url=cta_url_com_utm,
            preheader_text=preheader_text,
            apenas_produtos=apenas_produtos
        )
        
        print("→ Iniciando busca de produtos...")
        html_gerado = buscar_produtos(
            produtos_selecionados,
            template_para_produtos, 
            utm_source, 
            utm_campaign,
            cor_botao
        )
        
        print("✓ Email gerado com sucesso!")
        
        return jsonify({
            'success': True,
            'redirect': '/resultado',
            'html': html_gerado
        })
        
    except Exception as e:
        print(f"✗ ERRO CRÍTICO na rota /gerar: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            'success': False,
            'error': f'Erro ao gerar email: {str(e)}'
        }), 500

@app.route('/resultado')
def resultado():
    """
    Página de resultado que mostra o email gerado
    """
    # O HTML será passado via POST do frontend
    return render_template('resultado.html')

@app.route('/api/produtos', methods=['GET'])
def api_produtos():
    """
    Retorna todos os produtos em formato JSON para o SKU Consult
    """
    try:
        produtos = carregar_produtos_planilha()
        
        # Formata os produtos no formato esperado pelo SKU Consult
        produtos_formatados = []
        for p in produtos:
            produtos_formatados.append({
                'COD_PRODUTO': p.get('sku', ''),
                'NOMEPRODUTOECOMM': p.get('nome', ''),
                'COD_BARRAS': p.get('ean', ''),
                'URLECOMMERCEIMG': p.get('imagem', ''),
                'PRODUCTURL': p.get('url', '')
            })
        
        return jsonify({
            'success': True,
            'produtos': produtos_formatados,
            'total': len(produtos_formatados)
        })
    except Exception as e:
        print(f"Erro ao buscar produtos para API: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'produtos': []
        }), 500

@app.route('/api/buscar-imagem-por-sku', methods=['POST'])
def buscar_imagem_por_sku():
    """
    Busca a URL da imagem de um produto pelo SKU no banco de dados do Google Sheets
    """
    try:
        data = request.get_json()
        sku = data.get('sku', '').strip()
        
        if not sku:
            return jsonify({'success': False, 'error': 'SKU não fornecido'}), 400
        
        produtos = carregar_produtos_planilha()
        sku_lower = sku.lower()
        
        # Busca exata pelo SKU
        for produto in produtos:
            if produto['sku'].lower() == sku_lower:
                return jsonify({
                    'success': True,
                    'sku': produto['sku'],
                    'nome': produto['nome'],
                    'imagem': produto['imagem']
                })
        
        return jsonify({
            'success': False,
            'error': f'SKU {sku} não encontrado no banco de dados'
        }), 404
        
    except Exception as e:
        print(f"Erro ao buscar imagem por SKU: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/proxy-imagem', methods=['POST'])
def proxy_imagem():
    """
    Proxy para baixar imagens de URLs externas, contornando CORS.
    Baixa a imagem do lado do servidor e retorna como blob.
    """
    try:
        import requests
        
        data = request.get_json()
        url = data.get('url', '').strip()
        
        if not url:
            return jsonify({'success': False, 'error': 'URL não fornecida'}), 400
        
        # Baixa a imagem com timeout
        headers = {
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36'
        }
        
        response = requests.get(url, headers=headers, timeout=10, stream=True)
        response.raise_for_status()
        
        # Detecta o tipo de conteúdo
        content_type = response.headers.get('Content-Type', 'image/jpeg')
        
        # Retorna a imagem como bytes
        return send_file(
            io.BytesIO(response.content),
            mimetype=content_type,
            as_attachment=False
        )
        
    except requests.exceptions.Timeout:
        return jsonify({'success': False, 'error': 'Timeout ao baixar imagem'}), 408
    except requests.exceptions.RequestException as e:
        print(f"Erro ao baixar imagem via proxy: {e}")
        return jsonify({'success': False, 'error': f'Erro ao baixar imagem: {str(e)}'}), 500
    except Exception as e:
        print(f"Erro no proxy de imagem: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/buscar-e-baixar-imagem-produto', methods=['POST'])
def buscar_e_baixar_imagem_produto():
    """
    Busca a URL do produto pelo SKU e faz webscraping para extrair e baixar a imagem.
    Similar à lógica do gerador de emails.
    """
    try:
        import requests
        from bs4 import BeautifulSoup
        import json
        
        data = request.get_json()
        sku = data.get('sku', '').strip()
        
        if not sku:
            return jsonify({'success': False, 'error': 'SKU não fornecido'}), 400
        
        # 1. Busca a URL do produto pelo SKU
        produtos = carregar_produtos_planilha()
        produto_encontrado = None
        
        for produto in produtos:
            if produto['sku'].lower() == sku.lower():
                produto_encontrado = produto
                break
        
        if not produto_encontrado:
            return jsonify({'success': False, 'error': f'SKU {sku} não encontrado'}), 404
        
        url_produto = produto_encontrado.get('url', '').strip()
        
        if not url_produto:
            return jsonify({'success': False, 'error': 'URL do produto não disponível'}), 404
        
        print(f"→ Buscando imagem do produto: {url_produto}")
        
        # 2. Faz o webscraping da página do produto
        session = requests.Session()
        session.headers.update({
            "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
            "Connection": "keep-alive",
        })
        
        try:
            resp = session.get(url_produto, timeout=(10, 30))
            resp.raise_for_status()
        except Exception as e:
            print(f"✗ Falha ao acessar URL do produto: {e}")
            return jsonify({'success': False, 'error': f'Erro ao acessar página do produto: {str(e)}'}), 500
        
        soup = BeautifulSoup(resp.text, 'html.parser')
        
        # 3. Extrai a imagem usando a mesma lógica do gerador
        def _extract_apollo_state(script_text: str) -> dict | None:
            if not script_text:
                return None
            idx = script_text.find("window.APOLLO_STATE")
            if idx == -1:
                return None
            start = script_text.find("{", idx)
            if start == -1:
                return None
            braces = 0
            end = -1
            for i in range(start, len(script_text)):
                ch = script_text[i]
                if ch == "{":
                    braces += 1
                elif ch == "}":
                    braces -= 1
                    if braces == 0:
                        end = i + 1
                        break
            if end == -1:
                return None
            json_str = script_text[start:end]
            json_str = json_str.replace(": undefined", ": null").replace(":undefined", ": null")
            try:
                return json.loads(json_str)
            except Exception:
                try:
                    return json.loads(json_str.encode("utf-8", "ignore").decode("utf-8"))
                except Exception:
                    return None

        def _extract_image_from_html(soup_obj: BeautifulSoup, nome_produto: str | None = None) -> str:
            """Extrai a imagem do HTML (mesmo método do gerador)"""
            placeholder = "https://via.placeholder.com/120"
            
            selectors = [
                'div.product-image-gallery-active-image img[src]',
                'div.product-image-gallery-desktop-view img[src]',
                'div.product-image-gallery img[src]',
                'div.static-image-viewer-container img[src]',
                'img[class*="product-image"][src]',
                'img[src][alt]'
            ]
            candidates = []
            for sel in selectors:
                el = soup_obj.select_one(sel)
                if el and el.get('src'):
                    src = (el.get('src') or '').strip()
                    if src and not src.startswith('data:'):
                        candidates.append((src, el.get('alt') or ''))
            
            if candidates:
                if nome_produto:
                    for src, alt in candidates:
                        if nome_produto.lower() in alt.lower():
                            return src
                return candidates[0][0]
            
            og = soup_obj.find('meta', property='og:image') or soup_obj.find('meta', attrs={'name': 'og:image'})
            if og and og.get('content'):
                content = og.get('content').strip()
                if content:
                    return content
            
            link_img = soup_obj.find('link', rel='image_src')
            if link_img and link_img.get('href'):
                href = link_img.get('href').strip()
                if href:
                    return href
            
            any_img = soup_obj.find('img', src=True)
            if any_img and any_img.get('src'):
                return any_img.get('src').strip()
            
            return placeholder

        def _resolve_image_url(apollo: dict, image_ref) -> str:
            """Resolve URL da imagem via APOLLO_STATE"""
            placeholder = "https://via.placeholder.com/120"
            ref_key = None
            if isinstance(image_ref, str):
                ref_key = image_ref
            elif isinstance(image_ref, dict):
                ref_key = image_ref.get("__ref") or image_ref.get("id")
            if not ref_key:
                return placeholder
            file_obj = apollo.get(ref_key)
            if isinstance(file_obj, dict):
                return file_obj.get("url") or file_obj.get("src") or placeholder
            return placeholder
        
        # Tenta extrair do HTML primeiro
        nome_produto = produto_encontrado.get('nome', '')
        url_imagem = _extract_image_from_html(soup, nome_produto)
        
        # Se não encontrou no HTML ou é SVG/GIF, tenta APOLLO_STATE
        if not url_imagem or url_imagem.endswith(('.svg', '.gif')) or 'placeholder' in url_imagem:
            script_tag = soup.find('script', id='main-states')
            if script_tag:
                apollo = _extract_apollo_state(script_tag.text)
                if apollo:
                    # Procura o produto no APOLLO_STATE
                    for key, value in apollo.items():
                        if isinstance(key, str) and key.startswith("PublicViewerProduct:") and isinstance(value, dict):
                            url_imagem = _resolve_image_url(apollo, value.get("image"))
                            if url_imagem and 'placeholder' not in url_imagem:
                                break
        
        if not url_imagem or 'placeholder' in url_imagem:
            print(f"✗ Não foi possível extrair URL da imagem para SKU {sku}")
            return jsonify({'success': False, 'error': 'Imagem não encontrada na página do produto'}), 404
        
        print(f"✓ URL da imagem encontrada: {url_imagem}")
        
        # 4. Baixa a imagem
        headers_img = {
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36',
            'Referer': url_produto
        }
        
        try:
            img_response = session.get(url_imagem, headers=headers_img, timeout=(10, 30), stream=True)
            img_response.raise_for_status()
            
            content_type = img_response.headers.get('Content-Type', 'image/jpeg')
            
            if not content_type.startswith('image/'):
                return jsonify({'success': False, 'error': 'URL não retorna uma imagem válida'}), 400
            
            print(f"✓ Imagem baixada: {len(img_response.content)} bytes")
            
            # Retorna a imagem
            return send_file(
                io.BytesIO(img_response.content),
                mimetype=content_type,
                as_attachment=False
            )
            
        except Exception as e:
            print(f"✗ Erro ao baixar imagem: {e}")
            return jsonify({'success': False, 'error': f'Erro ao baixar imagem: {str(e)}'}), 500
        
    except Exception as e:
        print(f"✗ Erro no webscraping: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# --- ROTA PARA ORGANIZAR IMAGENS EM PASTAS ---
# --- ROTA PARA ORGANIZAR IMAGENS EM PASTAS ---
# --- ROTA PARA ORGANIZAR IMAGENS EM PASTAS ---
# --- ROTA PARA ORGANIZAR IMAGENS EM PASTAS ---
# --- ROTA PARA ORGANIZAR IMAGENS EM PASTAS ---
# --- ROTA PARA ORGANIZAR IMAGENS EM PASTAS ---

@app.route('/processar_imagens', methods=['POST'])
def processar_imagens():
    """
    Recebe um .xlsx e gera um .zip com estrutura de pastas:
    - Nível 1: Dinâmica (se a planilha tiver essa coluna)
    - Nível 2: Criativo (coluna informada no formulário)
    Observação: não baixa imagens; cria .keep para preservar diretórios vazios.
    """
    try:
        excel_file = request.files.get('excel_file')
        creative_column = (request.form.get('creative_column') or '').strip()

        if not excel_file or not creative_column:
            return jsonify({'success': False, 'error': 'Arquivo e coluna do criativo são obrigatórios'}), 400

        try:
            from openpyxl import load_workbook
        except Exception:
            return jsonify({
                'success': False,
                'error': 'Dependência ausente: instale openpyxl (pip install openpyxl)'
            }), 500

        # Lê o Excel em memória
        data = excel_file.read()
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active

        # Cabeçalhos
        headers = [str(c.value or '').strip() for c in ws[1]]

        def _norm(s: str) -> str:
            s = str(s or '').strip().lower()
            s = unicodedata.normalize('NFKD', s)
            return ''.join(ch for ch in s if not unicodedata.combining(ch))

        header_norm_map = {_norm(h): h for h in headers}

        # Resolve coluna do criativo (nome fornecido pelo usuário)
        col_creative_name = header_norm_map.get(_norm(creative_column))
        if not col_creative_name:
            # tenta correspondência parcial
            col_creative_name = next((orig for norm, orig in header_norm_map.items() if _norm(creative_column) in norm), None)
        if not col_creative_name:
            return jsonify({'success': False, 'error': f'Coluna do Criativo não encontrada: "{creative_column}"'}), 400

        # Tenta detectar coluna "Dinâmica"
        dinamica_aliases = ['dinamica', 'dinâmica', 'categoria', 'campanha', 'grupo', 'etiqueta', 'pasta']
        col_dinamica_name = None
        for alias in dinamica_aliases:
            if alias in header_norm_map:
                col_dinamica_name = header_norm_map[alias]
                break

        idx_map = {h: i for i, h in enumerate(headers)}
        idx_creative = idx_map[col_creative_name]
        idx_dinamica = idx_map.get(col_dinamica_name) if col_dinamica_name else None

        # Diretório temporário base
        base_dir = tempfile.mkdtemp(prefix='temp_organizador_')

        def safe(name: str) -> str:
            name = str(name or '').strip()
            if not name:
                return ''
            name = unicodedata.normalize('NFKD', name)
            name = ''.join(ch for ch in name if not unicodedata.combining(ch))
            name = name.replace('/', '-')
            name = re.sub(r'[^A-Za-z0-9\-\._ ]+', '', name)
            name = re.sub(r'\s+', ' ', name).strip()
            return name[:120]  # evita nomes muito longos

        # Cria a estrutura de diretórios
        for row in ws.iter_rows(min_row=2, values_only=True):
            creative_val = safe(row[idx_creative] if idx_creative is not None else '')
            if not creative_val:
                continue

            # Dinâmica em MAIÚSCULO (nível 1)
            dinamica_val = (safe(row[idx_dinamica] if idx_dinamica is not None else '') or '').upper()
            nivel1 = os.path.join(base_dir, dinamica_val) if dinamica_val else base_dir
            final_path = os.path.join(nivel1, creative_val)

            os.makedirs(final_path, exist_ok=True)
            keep_file = os.path.join(final_path, '.keep')
            if not os.path.exists(keep_file):
                with open(keep_file, 'w', encoding='utf-8') as f:
                    f.write('')

        # Gera o ZIP
        # Define nome do ZIP a partir do nome da planilha enviada
        orig_name = os.path.splitext(os.path.basename(excel_file.filename or 'pastas'))[0]
        zip_name_base = safe(orig_name) or 'pastas'

        # Cria o ZIP fora do diretório que será compactado para evitar "zip dentro do zip"
        zip_out_dir = tempfile.mkdtemp(prefix='zip_out_')
        zip_base = os.path.join(zip_out_dir, zip_name_base)
        zip_path = shutil.make_archive(zip_base, 'zip', root_dir=base_dir)

        # Retorna o arquivo para download
        return send_from_directory(
            os.path.dirname(zip_path),
            os.path.basename(zip_path),
            as_attachment=True,
            download_name=f'{zip_name_base}.zip'
        )

    except Exception as e:
        print(f'Erro em /processar_imagens: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

# --- CARREGAMENTO INICIAL DO CACHE ---
print("\n" + "="*60)
print("🚀 Iniciando Email Marketing Generator...")
print("="*60)

# Carrega produtos no cache ao iniciar o app
print("\n→ Pré-carregando catálogo de produtos...")
produtos_iniciais = carregar_produtos_planilha()
if produtos_iniciais:
    print(f"✓ Cache inicial criado com {len(produtos_iniciais)} produtos")
    if produtos_iniciais:
        print(f"  Exemplo: {produtos_iniciais[0]['nome'][:60]}...")
else:
    print("⚠ Nenhum produto carregado no cache inicial")

print("\n" + "="*60)
print("✓ App pronto para uso!")
print("="*60 + "\n")

# --- Inicia o servidor ---
if __name__ == '__main__':
   app.run(debug=True)